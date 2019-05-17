"""
Author: Dan Ollerenshaw

Simple python demonstrations. Code here was written fairly quickly, this 
shouldn't necessarily be regarded as "best practice".

You should be able to run this whole thing on any laptop, all data is generated
by the code. Be aware that it will save some data on your D drive.

Note: the "#%%" lines are Spyder's way of mimicking the "cell" structure that 
you see in Jupyter Notebooks.
They allow you to run a "block" of code, where each blocked is delimited by #%%

You can still run line by line, or the entire script if you wish.

Demos:
(1) making .png charts
(2) editing an excel file to create an excel chart
(3) "chunking" a big dataset
(4) misc usefulness
(5) running a regression
(6) pyautogui fun

"""

# import the modules to use
import calendar
import datetime as dt
import getpass
import matplotlib.pyplot as plt
import numpy as np
import os
import pandas as pd
import statsmodels.formula.api as sm
import time

from openpyxl.chart import LineChart, Reference, Series
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

"""
FYI: these modules may need installing separately (they don't come with anaconda)
to install new python modules at ONS, please follow the 
"installing python packages with pip" word document
pinned here: https://www.yammer.com/ons.gsi.gov.uk/#/files/120252827
if you get stuck on this, you can email me (daniel.ollerenshaw@ons.gov.uk) for help.
"""
import openpyxl
import pyautogui

#%%
#==============================================================================
# (1) Making plots
#==============================================================================

# setup: create some dummy data. random values for each month.
months = [m.lower().capitalize()[:3] for m in calendar.month_name[1:]]
month_dict = {m:i for i,m in enumerate(months)}

npoints = 100
dummy_data = pd.concat(
                [pd.DataFrame({'month':[m]*npoints,
                               'value':np.random.normal(size=npoints)})
                for m in months]
            )

# line graph of the sums each month:   
totals = dummy_data.groupby('month').sum().reset_index()
totals['order'] = totals['month'].map(month_dict)
totals = totals.sort_values(by='order').set_index('month')[['value']]

# make a folder to save outputs to:
try:
    os.mkdir('D:/demo/')
except FileExistsError:
    pass

plt.plot(totals['value'].values)
plt.xlabel('month')
plt.ylabel('total')
plt.title('example line chart!', fontsize=16);
plt.xticks(range(totals.shape[0]), totals.index) # add labels
plt.savefig('D:/demo/line.png')
plt.close()

# histogram for a particular month
july = dummy_data[dummy_data['month'] == 'Jul']

plt.hist(july['value'],
         edgecolor='black',
         bins=20)
plt.xlabel('Value')
plt.ylabel('Count')
plt.title('example histogram!', fontsize=16)
plt.savefig('D:/demo/hist.png')
plt.close()

# cumulative frequency plot 
plt.step(sorted(july['value']), np.arange(july.shape[0]))
plt.axvline(0, linestyle='--', color='red', alpha=0.5)
plt.title('example cumulative frequency, with a line at zero', fontsize=16)
plt.savefig('D:/demo/cumfreq.png')
plt.close()

#%%
#==============================================================================
# (2) Making an excel plot using the openpyxl module
# The code is a bit messy for this, it's probably only worth doing this if you
# do the same thing manually over and over again!

# This code assumes that you've just downloaded the latest construction dataset:
# https://www.ons.gov.uk/businessindustryandtrade/constructionindustry/datasets/outputintheconstructionindustry
#==============================================================================
# change working directory to downloads folder
os.chdir('C:/Users/{}/Downloads/'.format(getpass.getuser()))

# find all the excel files in the downloads folder
excel_files = [f for f in os.listdir() if f.endswith('.xlsx')]
# select the one that got changed last (hopefully the one you just downloaded)
latest_excel_file = max([f for f in excel_files], key=os.path.getctime)

# open the file with the openpyxl module
workbook = openpyxl.load_workbook(latest_excel_file)

#check file is as expected, and throw an error if sheetnames don't match this list
expected_sheets = ['Cover Sheet',
                   'Contents',
                   'Table 1a',
                   'Table 1b',
                   'Table 2a',
                   'Table 2b',
                   'Table 3a',
                   'Table 3b',
                   'Table 3c',
                   'Table 3d',
                   'Table 4',
                   'Table 4a',
                   'Table 5',
                   'Table 6',
                   'Table 7a',
                   'Table 7b',
                   'Table 9a',
                   'Table 10',
                   'Table 11',
                   'Table 12',
                   'Table 13',
                   'Table 14']

assert workbook.get_sheet_names() == expected_sheets, "Error, sheetnames don't match!"

# open a particular sheet
volm = workbook.get_sheet_by_name('Table 2a')

# Now create the new entries: some simple calculations
# New cell entries are coloured yellow for clarity

# set-up
yellow = PatternFill('solid','ffff00') # set colour for python created cells. See http://www.discoveryplayground.com/computer-programming-for-kids/rgb-colors/

last_row = volm.max_row
last_col = volm.max_column

# There is a case where these don't work - if the creator of the spreadsheet
# has left the cursor highlighting empty cells below.
# This needs to be fixed case-by-case, can check for None values

# insert first calc
volm.cell(row=last_row+2,column=2).value = 'M/M growth'
volm.cell(row=last_row+2,column=2).fill = yellow # could do this all in one at the end
    
for i in volm['C'+str(last_row+2):'P'+str(last_row+2)][0]: # probably a better way to do this
    i.value = '='+i.coordinate[0]+str(last_row)+'/'+i.coordinate[0]+str(last_row-1)+'-1' # percentage change formula
    i.fill = yellow
    i.number_format = '0.0%' # format as percentage

# Insert second calc
volm.cell(row=last_row+3,column=2).value = 'M/M cont'
volm.cell(row=last_row+3,column=2).fill = yellow

for i in volm['C'+str(last_row+3):'P'+str(last_row+3)][0]:
    i.value = '=100*(('+i.coordinate[0]+str(last_row)+'-'+i.coordinate[0]+str(last_row-1)+')/'+'$P'+str(last_row-1)+')'
    i.fill = yellow
    i.number_format = '0.0'

# Insert a chart in a different sheet
voli = workbook.get_sheet_by_name('Table 1a')

last_row = voli.max_row
last_col = voli.max_column

c1 = LineChart() # create a blank line chart


def get_month_start():
    """Search for the '2010' and 'Jan' adjacent cells and get the row ref
    """
    for i in range(1,last_row+1): 
        for j in range(1,last_col+1):
            if (volm.cell(row=i,column=j).value == 2010) and (volm.cell(row=i,column=j+1).value == 'Jan '): # yeah there's a space there...
                return int((volm.cell(row=i,column=j).coordinate)[1:])


month_data_start = get_month_start()


def get_relevant_cols():
    """Search for the cols with private new housing and all work
    """
    for i in range(1,last_row+1): 
        for j in range(1,last_col+1):
            if (voli.cell(row=i,column=j).value == 'Private') and (voli.cell(row=i-1,column=j-1).value == 'New Housing'):
                p = (volm.cell(row=i,column=j).coordinate)[0]
            if (voli.cell(row=i,column=j).value == 'All Work'):
                a = (volm.cell(row=i,column=j).coordinate)[0]
    return p, a


pnh_col_ref = column_index_from_string(get_relevant_cols()[0])
aw_col_ref = column_index_from_string(get_relevant_cols()[1])

P_N_H = Reference(voli,min_row=month_data_start, max_row=last_row, min_col=pnh_col_ref, max_col=pnh_col_ref) # define range to draw data from
P_N_H_plus = Series(P_N_H, title='Private new housing') # add legend label
all_work = Reference(voli,min_row=month_data_start, max_row=last_row, min_col=aw_col_ref, max_col=aw_col_ref)
all_work_plus = Series(all_work,title='All Work')

c1.append(P_N_H_plus) # add data to chart
c1.append(all_work_plus)


# Format chart. Documentation is a bit scant here.
labs = Reference(voli,min_row=month_data_start,max_row=last_row,min_col=1,max_col=2)
c1.set_categories(labs)

c1.legend.position=('t')
# would be good to remove this hardcoding!
c1.y_axis.scaling.min=40
c1.y_axis.scaling.max=150

# ONS colours, http://www.rapidtables.com/convert/color/rgb-to-hex.htm
ONS_colours = ['274796','F5942F','E73F40','7BCAE2','979796','E9E117'] # in order of use
# new based on change on 24/05/2017
ONS_new_colours = ['0075A3','E2BC22','234D70','36ADD9','266D4A','0064B9','AAD6E8','112941','009FE3','5BC5F2']

s1 = c1.series[0]
s1.graphicalProperties.line.solidFill = ONS_colours[0] 
s2 = c1.series[1]
s2.graphicalProperties.line.solidFill = ONS_colours[1]

voli.add_chart(c1,'C'+str(int(last_row+3)))

workbook.save('D:/python_construction_file.xlsx')
print('New file saved!')

#%%
#==============================================================================
# (3) "chunking" a big dataset
#==============================================================================

# Suppose you have a dataset that is too big to read into memory. What do you
# do?

# Create some dummy data:
size = 10**5
df = pd.DataFrame({'city':np.random.choice(['cardiff','newport','swansea'],size=size),
                   'value':np.random.randint(0,100, size=size)})

# Suppose that this data was too big to read in one go, so you want to save
# three separate, smaller files, one for each of "cardiff, newport, swansea":
    
df.to_csv('D:/big_dataset.csv', index=None)

# we read the data in "chunks" and save out each chunk bit by bit:

chunksize = 500
# this doesn't actually read anything, just sets it up
chunks = pd.read_csv('D:/big_dataset.csv', chunksize=chunksize)

total_rows_written = 0
for n,chunk in enumerate(chunks):
    # read a chunk of rows
    df = chunk
    # decide whether to output the column names or not
    header = True if n == 0 else False
    # split chunk by the id_var and write out to 3 separate files
    for id_var, dataset in df.groupby('city'):
        dataset.to_csv('D:/dummy_output_{}.csv'.format(id_var),
                     mode='a', # a for append to existing file
                     index=False,
                     header=header)
    total_rows_written += chunksize
    print('Written out {} rows from data.'.format(total_rows_written))
print('Finished writing out data!')

#%%
#==============================================================================
# (4) Miscellaneous usefulness. E.g. for one project I needed to get a list
# of all the second Tuesdays of the month! 
#==============================================================================

def find_second_tuesday(year, month, give_tol=True):
    """ Returns a datetime object of the 2nd Tuesday of the month.
    
        Enter year/month as integers.
        
        give_tol param:
            If True, return the second tuesday and one day either side as a list
            If False, just return the second tuesday.
    """
    cal = calendar.Calendar()
    month_cal = cal.monthdays2calendar(year, month)
    check = month_cal[0][1]
    if check[0] == 0:
        second_week = month_cal[2]
    else:
        second_week = month_cal[1]
    second_tues = second_week[1][0]
    
    if give_tol:
        return [dt.datetime(year, month, d) for d in [second_tues-1, second_tues, second_tues+1]]
    else:
        return dt.datetime(year, month, second_tues)
    
# example use:
find_second_tuesday(2018, 10, give_tol=False)

#%%
#==============================================================================
# (5) running a regression
#==============================================================================

# create some data:
ix = []
for year in range(2000,2018):
    for q in range(1,5):
        quarter = '{}Q{}'.format(year,q)
        ix.append(quarter)

# these are real ONS data (slightly out of date now)
gdp_values = [351979,354546,355557,356094,360749,363235,365696,367157,368690,371397,374205,377469,380565,384014,387841,391048,393247,395113,395867,398259,400558,404875,409270,414816,416183,417146,417741,419236,423472,426600,429845,433079,433710,430887,423618,414037,407291,406353,406641,408298,410542,414699,417024,417507,419828,420109,421918,422965,424834,424468,429319,428321,431025,433271,436560,438707,442425,446519,450141,453891,455026,457214,458456,461622,462333,465113,467437,470527,471524]

unemployment_values = [1682,1599,1548,1520,1481,1468,1487,1520,1515,1520,1566,1516,1531,1467,1502,1457,1434,1437,1401,1423,1416,1437,1437,1567,1604,1685,1693,1701,1704,1658,1650,1608,1622,1680,1840,2003,2235,2448,2475,2453,2526,2488,2470,2503,2483,2540,2664,2684,2633,2582,2538,2535,2541,2515,2481,2358,2213,2061,1957,1872,1830,1848,1752,1692,1693,1640,1606,1594,1541]

df = pd.DataFrame({'GDP':gdp_values,
                   'unemployment':unemployment_values},
                    index=ix[:len(gdp_values)])

# We're not trying for a serious model here, but lets take differences anyway

df['dGDP'] = df['GDP'].diff(periods=1)

df['dU'] = df['unemployment'].diff(periods=1)

df['dU_lagged'] = df['dU'].shift(1)

# Run an OLS regression with dGDP as the dependent variable
# and a 1-period lag of dU as the independent variable

# There are lots of ways we can do this with python.

results2 = sm.ols(formula='dGDP~dU_lagged', data=df).fit()

print(results2.summary())

#%%
#==============================================================================
# (6) Fun with pyautogui!
# There are some practical applications for this...
#==============================================================================

# from https://pyautogui.readthedocs.io/en/latest/introduction.html
# switch to MS paint just after running this and position cursor in the middle
time.sleep(3)
distance = 200
while distance > 0:
    pyautogui.dragRel(distance, 0, duration=0.5)   # move right
    distance -= 5
    pyautogui.dragRel(0, distance, duration=0.5)   # move down
    pyautogui.dragRel(-distance, 0, duration=0.5)  # move left
    distance -= 5
    pyautogui.dragRel(0, -distance, duration=0.5)  # move up







# These are just a few examples... the possibilities are endless!